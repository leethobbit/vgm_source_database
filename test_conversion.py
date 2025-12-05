"""Test conversion with logging."""
import sys
import traceback
from pathlib import Path

# Redirect output to file
log_file = open('conversion_log.txt', 'w', encoding='utf-8')
sys.stdout = log_file
sys.stderr = log_file

try:
    print("=" * 60)
    print("CONVERSION TEST")
    print("=" * 60)
    
    import openpyxl
    print("✓ openpyxl imported")
    
    import yaml
    print("✓ yaml imported")
    
    excel_file = Path("dev_reference/NEWER VGM Sound Sources.xlsx")
    print(f"\nChecking file: {excel_file}")
    print(f"Exists: {excel_file.exists()}")
    print(f"Absolute: {excel_file.absolute()}")
    
    if not excel_file.exists():
        print("ERROR: File not found!")
        sys.exit(1)
    
    print("\nLoading workbook...")
    wb = openpyxl.load_workbook(str(excel_file), data_only=True)
    print(f"✓ Loaded workbook with {len(wb.sheetnames)} sheets")
    print(f"Sheets: {wb.sheetnames[:5]}")
    
    # Skip Rules
    game_sheets = [name for name in wb.sheetnames if name.lower() != "rules"]
    print(f"\nGame sheets to process: {len(game_sheets)}")
    print(f"First few: {game_sheets[:3]}")
    
    # Test first sheet
    if game_sheets:
        sheet_name = game_sheets[0]
        print(f"\nTesting sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Get headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                break
        print(f"Headers ({len(headers)}): {headers[:5]}")
        print(f"Total rows: {ws.max_row}")
        
        # Check a few rows
        print("\nFirst 5 data rows:")
        for row_idx in range(2, min(7, ws.max_row + 1)):
            row_data = {}
            for col_idx, header in enumerate(headers[:3], 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    row_data[header] = str(cell.value).strip()[:50]
            if row_data:
                print(f"  Row {row_idx}: {row_data}")
    
    print("\n✓ Basic test passed")
    print("\nNow trying full conversion...")
    
    from convert_excel_to_yaml import ExcelConverter
    converter = ExcelConverter(str(excel_file), "fixtures")
    converter.convert()
    
    print("\n✓ Conversion complete!")
    
except Exception as e:
    print(f"\nERROR: {e}")
    traceback.print_exc()
finally:
    log_file.close()
    print("\nLog written to conversion_log.txt")
