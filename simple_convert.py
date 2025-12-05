"""Simple Excel to YAML converter with file-based output."""
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import openpyxl
    import yaml
except ImportError:
    print("ERROR: Install openpyxl and pyyaml")
    sys.exit(1)

def slugify(text):
    """Convert text to slug."""
    text = text.lower().strip()
    text = re.sub(r"[^\w\s-]", "", text)
    text = re.sub(r"[-\s]+", "-", text)
    return text

def is_game_title(first_value):
    """Check if a value looks like a game title."""
    if not first_value or len(first_value) < 10:
        return False
    
    # Skip section headers
    section_patterns = ["STUFF TO FIND", "SOURCES", "NOTES", "DESCRIPTION", "SECTION"]
    if any(pattern in first_value.upper() for pattern in section_patterns):
        return False
    
    # Game indicators
    indicators = [":", "(", "HD", "Remaster", "Remake", "Edition", "Wii", "PlayStation", "Xbox", "Nintendo", "PC", "Switch"]
    if any(indicator in first_value for indicator in indicators):
        return True
    
    # Long standalone text
    if len(first_value) >= 20:
        return True
    
    return False

# Initialize
excel_file = Path("dev_reference/NEWER VGM Sound Sources.xlsx")
output_dir = Path("fixtures")
output_dir.mkdir(exist_ok=True)

log = []
log.append("Starting conversion...")
log.append(f"Excel file: {excel_file}")
log.append(f"Exists: {excel_file.exists()}")

if not excel_file.exists():
    log.append("ERROR: File not found!")
    with open("conversion_error.txt", "w") as f:
        f.write("\n".join(log))
    sys.exit(1)

try:
    wb = openpyxl.load_workbook(str(excel_file), data_only=True)
    log.append(f"Loaded workbook with {len(wb.sheetnames)} sheets")
    
    game_sheets = [name for name in wb.sheetnames if name.lower() != "rules"]
    log.append(f"Game sheets: {len(game_sheets)}")
    
    # Data structures
    gametags = {}
    companies = {}
    products = {}
    banks = {}
    games = {}
    sound_sources = []
    skipped_lines = []
    
    gametag_pk = 1
    company_pk = 1
    product_pk = 1
    bank_pk = 1
    game_pk = 1
    soundsource_pk = 1
    
    # Process each sheet
    for sheet_name in game_sheets:
        log.append(f"\nProcessing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        # Create GameTag
        if sheet_name not in gametags:
            gametags[sheet_name] = {
                "pk": gametag_pk,
                "slug": slugify(sheet_name),
            }
            gametag_pk += 1
        
        # Get headers
        headers = []
        for cell in ws[1]:
            if cell.value:
                headers.append(str(cell.value).strip())
            else:
                break
        
        if not headers:
            continue
        
        current_game_title = None
        current_game_pk = None
        
        # Process rows
        for row_idx in range(2, ws.max_row + 1):
            # Get row data
            row_data = {}
            first_value = None
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    value = str(cell.value).strip()
                    row_data[header] = value
                    if col_idx == 1:
                        first_value = value
            
            # Skip empty rows
            if not first_value:
                continue
            
            # Check if game title
            if is_game_title(first_value):
                # Save previous game
                if current_game_title:
                    games[current_game_title] = current_game_pk
                
                # New game
                current_game_title = first_value
                if current_game_title not in games:
                    current_game_pk = game_pk
                    games[current_game_title] = game_pk
                    game_pk += 1
                else:
                    current_game_pk = games[current_game_title]
                
                log.append(f"  Row {row_idx}: Game title: {current_game_title}")
            
            elif current_game_title:
                # Try to parse as sound source
                # Skip section headers
                if any(p in first_value.upper() for p in ["STUFF TO FIND", "SOURCES", "NOTES", "DESCRIPTION"]):
                    skipped_lines.append({
                        "sheet": sheet_name,
                        "row": row_idx,
                        "content": first_value[:100],
                        "reason": "Section header"
                    })
                    continue
                
                # Create sound source
                ss_name = first_value
                ss_bank = row_data.get("Bank", "").strip() or None
                ss_product = row_data.get("Product", "").strip() or None
                ss_company = row_data.get("Company", "").strip() or None
                ss_notes = row_data.get("Notes", "").strip() or row_data.get("Description", "").strip() or ""
                
                # Create company/product/bank if needed
                company_pk_val = None
                product_pk_val = None
                bank_pk_val = None
                
                if ss_company:
                    if ss_company not in companies:
                        companies[ss_company] = company_pk
                        company_pk += 1
                    company_pk_val = companies[ss_company]
                
                if ss_product and ss_company:
                    key = (ss_product, ss_company)
                    if key not in products:
                        products[key] = product_pk
                        product_pk += 1
                    product_pk_val = products[key]
                
                if ss_bank and ss_product and ss_company:
                    key = (ss_bank, ss_product, ss_company)
                    if key not in banks:
                        banks[key] = bank_pk
                        bank_pk += 1
                    bank_pk_val = banks[key]
                
                # Must have at least product or bank
                if not bank_pk_val and not product_pk_val:
                    # Try to create product from company
                    if ss_company:
                        key = (ss_company, ss_company)  # Use company as product
                        if key not in products:
                            products[key] = product_pk
                            product_pk += 1
                        product_pk_val = products[key]
                
                if bank_pk_val or product_pk_val:
                    sound_sources.append({
                        "name": ss_name,
                        "bank": bank_pk_val,
                        "product": product_pk_val,
                        "game_pk": current_game_pk,
                        "notes": ss_notes,
                    })
                else:
                    skipped_lines.append({
                        "sheet": sheet_name,
                        "row": row_idx,
                        "content": first_value[:100],
                        "reason": "No bank or product"
                    })
        
        # Save last game
        if current_game_title:
            games[current_game_title] = current_game_pk
    
    log.append(f"\nSummary:")
    log.append(f"  GameTags: {len(gametags)}")
    log.append(f"  Companies: {len(companies)}")
    log.append(f"  Products: {len(products)}")
    log.append(f"  Banks: {len(banks)}")
    log.append(f"  Games: {len(games)}")
    log.append(f"  SoundSources: {len(sound_sources)}")
    log.append(f"  Skipped: {len(skipped_lines)}")
    
    # Generate YAML files
    log.append("\nGenerating YAML files...")
    
    # 1. GameTags
    gametags_yaml = []
    for name, data in sorted(gametags.items()):
        gametags_yaml.append({
            "model": "vgm_source_database.games.GameTag",
            "pk": data["pk"],
            "fields": {
                "name": name,
                "slug": data["slug"],
                "description": "",
            },
        })
    
    # 2. Companies
    companies_yaml = []
    for name, pk in sorted(companies.items()):
        companies_yaml.append({
            "model": "vgm_source_database.sources.Company",
            "pk": pk,
            "fields": {
                "name": name,
                "notes": "",
            },
        })
    
    # 3. Products
    products_yaml = []
    for (name, company_name), pk in sorted(products.items()):
        company_pk_val = companies.get(company_name)
        if company_pk_val:
            products_yaml.append({
                "model": "vgm_source_database.sources.Product",
                "pk": pk,
                "fields": {
                    "name": name,
                    "company": company_pk_val,
                    "notes": "",
                },
            })
    
    # 4. Banks
    banks_yaml = []
    for (name, product_name, company_name), pk in sorted(banks.items()):
        product_key = (product_name, company_name)
        product_pk_val = products.get(product_key)
        if product_pk_val:
            banks_yaml.append({
                "model": "vgm_source_database.sources.Bank",
                "pk": pk,
                "fields": {
                    "name": name,
                    "product": product_pk_val,
                    "notes": "",
                },
            })
    
    # 5. Games (need to assign tags)
    games_yaml = []
    for title, pk in sorted(games.items()):
        # Find which tag this game belongs to
        tag_pks = []
        for sheet_name, tag_data in gametags.items():
            # This is simplified - in reality we'd track which sheet each game came from
            # For now, assign to first tag (we'll need to improve this)
            pass
        
        # Extract release year
        release_year = None
        year_match = re.search(r"\((\d{4})\)", title)
        if year_match:
            try:
                release_year = int(year_match.group(1))
            except ValueError:
                pass
        
        games_yaml.append({
            "model": "vgm_source_database.games.Game",
            "pk": pk,
            "fields": {
                "title": title,
                "release_date": None,
                "release_year": release_year,
                "album_artists": [],
                "tags": [],  # Will need to fix this
                "notes": "",
            },
        })
    
    # 6. SoundSources
    soundsources_yaml = []
    for ss in sound_sources:
        soundsources_yaml.append({
            "model": "vgm_source_database.sources.SoundSource",
            "pk": soundsource_pk,
            "fields": {
                "name": ss["name"],
                "bank": ss["bank"],
                "product": ss["product"],
                "discoverers": [],
                "games": [ss["game_pk"]] if ss["game_pk"] else [],
                "songs": [],
                "notes": ss["notes"] or "",
            },
        })
        soundsource_pk += 1
    
    # Write files
    files_to_write = [
        ("games_gametags.yaml", gametags_yaml),
        ("sources_companies.yaml", companies_yaml),
        ("sources_products.yaml", products_yaml),
        ("sources_banks.yaml", banks_yaml),
        ("games_games.yaml", games_yaml),
        ("sources_soundsources.yaml", soundsources_yaml),
    ]
    
    for filename, data in files_to_write:
        filepath = output_dir / filename
        with open(filepath, "w", encoding="utf-8") as f:
            yaml.dump(data, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
        log.append(f"  Generated: {filename} ({len(data)} entries)")
    
    # Write skipped lines report
    if skipped_lines:
        report_path = output_dir / "skipped_lines_report.txt"
        with open(report_path, "w", encoding="utf-8") as f:
            f.write("SKIPPED LINES REPORT\n")
            f.write("=" * 60 + "\n\n")
            f.write(f"Total skipped lines: {len(skipped_lines)}\n\n")
            
            by_sheet = defaultdict(list)
            for item in skipped_lines:
                by_sheet[item["sheet"]].append(item)
            
            for sheet_name, items in sorted(by_sheet.items()):
                f.write(f"\nSheet: {sheet_name}\n")
                f.write("-" * 60 + "\n")
                for item in items:
                    f.write(f"  Row {item['row']}: {item['content']}\n")
                    f.write(f"    Reason: {item['reason']}\n")
        
        log.append(f"  Skipped lines report: {report_path}")
    
    log.append("\nConversion complete!")
    
except Exception as e:
    log.append(f"\nERROR: {e}")
    import traceback
    log.append(traceback.format_exc())

# Write log
with open("conversion_log.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(log))

print("\n".join(log))
