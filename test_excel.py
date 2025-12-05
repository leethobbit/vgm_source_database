"""Simple test to examine Excel file."""
import openpyxl
import json

wb = openpyxl.load_workbook('dev_reference/NEWER VGM Sound Sources.xlsx', data_only=True)

result = {
    'sheet_names': wb.sheetnames,
    'num_sheets': len(wb.sheetnames),
    'game_sheets': wb.sheetnames[1:],
    'first_game_sheet': {}
}

# Examine first game sheet
if len(wb.sheetnames) > 1:
    sheet_name = wb.sheetnames[1]
    ws = wb[sheet_name]
    
    # Get headers
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
        else:
            break
    
    result['first_game_sheet'] = {
        'name': sheet_name,
        'headers': headers,
        'num_headers': len(headers),
        'sample_rows': []
    }
    
    # Get first 2 data rows
    for row_idx in range(2, min(4, ws.max_row + 1)):
        row_data = {}
        for col_idx, header in enumerate(headers, 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_data[header] = str(cell_value)[:100]  # Truncate long values
        if row_data:
            result['first_game_sheet']['sample_rows'].append(row_data)

# Write to file
with open('excel_analysis.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, indent=2, ensure_ascii=False)

print("Analysis complete. Check excel_analysis.json")
